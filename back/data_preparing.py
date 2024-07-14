import pandas as pd
import numpy as np
import os
import openpyxl


def getting_ld_gr(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    ld_number = None
    academic_group = None

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Номер ЛД" or cell.value == "Личное дело":
                ld_number_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                ld_number = ld_number_cell.value
            elif cell.value == "Учебная группа":
                academic_group_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                academic_group = academic_group_cell.value


    return ld_number, academic_group


def preparing_data(file_path):
    try:
        df = pd.read_excel(file_path, engine='openpyxl')

    except Exception as e:
        raise ValueError(f"Ошибка при чтении Excel файла: {e}")

    grade_mapping = {
        "Отлично": 5,
        "Хорошо": 4,
        "Удовлетворительно": 3,
        "Неудовлетворительно": 2,
        "зачтено": 5,
        "не зачтено": 2,
        "Неявка": 2,
        "Неявка по ув.причине": np.nan,
        "Не допущен": 2
    }

    df = df[df['Учебный год'] != '2023 - 2024']
    df = df[df['Учебный год'] != '2024 - 2025']
    df = df[df['Учебный год'] != '2025 - 2026']

    df['Оценка (успеваемость)'] = df['Оценка (успеваемость)'].replace(grade_mapping)
    df['Оценка (успеваемость)'] = df['Оценка (успеваемость)'].fillna('not given')  # Замена пропусков на not given

    df['Учебный год начало'] = df['Учебный год'].str.split('-').str[0].astype(int)
    df['Курс'] = df['Учебный год начало'] - (df['Учебная группа'].str.extract(r'(\d{2})')[0].astype(int) + 2000)

    mask = (df['Курс'] >= 3) & (df['Оценка (успеваемость)'] == 'not given')
    indices_to_drop = df[mask].sample(frac=0.2, random_state=42).index
    df.drop(indices_to_drop, inplace=True)

    df['Оценка (успеваемость)'].replace('not given', 2, inplace=True)

    df['Полугодие номер'] = df['Полугодие'].replace({'I полугодие': 1, 'II полугодие': 2})
    df['Семестр код'] = df['Учебный год начало'] * 10 + df['Полугодие номер']
    last_semester = df.groupby('hash')['Семестр код'].transform("max")

    df_last_semester = df[df['Семестр код'] == last_semester]
    df_previous_semesters = df[df['Семестр код'] != last_semester]

    grouped_prev = df_previous_semesters.groupby('hash')['Оценка (успеваемость)']
    dol_5 = grouped_prev.apply(lambda x: (x == 5).mean())
    dol_4 = grouped_prev.apply(lambda x: (x == 4).mean())
    dol_3 = grouped_prev.apply(lambda x: (x == 3).mean())

    # Агрегирование данных по двойкам для последнего семестра
    dol_2_last = df_last_semester.groupby('hash')['Оценка (успеваемость)'].apply(lambda x: (x == 2).mean())

    # Объединение результатов
    final_dataset = pd.DataFrame({
        'hash': dol_5.index,
        'Доля 5': dol_5,
        'Доля 4': dol_4,
        'Доля 3': dol_3,
        'Доля 2 последний семестр': dol_2_last.reindex(dol_5.index, fill_value=0)
    })
    final_dataset = final_dataset.drop(columns=['hash'])

    final_dataset = final_dataset.round(2)
    final_dataset = final_dataset.applymap(lambda x: int(x * 100))
    csv_path = 'dataset.csv'
    final_dataset.to_csv(csv_path, index=False)

    return csv_path

